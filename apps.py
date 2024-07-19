import streamlit as st
import camelot
import pandas as pd
import re
import pdfplumber
from openpyxl import Workbook
import io
import tempfile

def extract_xyz_bank(file):
    # Save the uploaded file to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
        tmp_file.write(file.read())
        temp_file_path = tmp_file.name
    
    # Use the temporary file path with camelot
    tables = camelot.read_pdf(temp_file_path, flavor="stream", pages="all", row_tol=15, strip_text='\n')
    
    # Initialize an empty list to store DataFrames from each page
    dfs = []
    
    for table in tables:
        tempDf = table.df
        if len(tempDf) > 1:  # Check if the table has more than one row
            columns_list = tempDf.iloc[1]
            df = tempDf.rename(columns=columns_list).drop(tempDf.index[0]).reset_index(drop=True)
            df = df.drop(df.index[0]).reset_index(drop=True)
            if 'Post Date' in df.columns and 'Value Date' in df.columns:
                df = df[['Post Date','Value Date','Particular','Debit','Credit','Balance']]
                dfs.append(df)
    
    # Concatenate all DataFrames
    if dfs:
        final_df = pd.concat(dfs, ignore_index=True)
        return final_df
    else:
        return pd.DataFrame()

def extract_yzx_bank(file):
    # Save the uploaded file to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
        tmp_file.write(file.read())
        temp_file_path = tmp_file.name

    # Use the temporary file path with camelot
    tables = camelot.read_pdf(temp_file_path, flavor="stream", pages="all", row_tol=5, strip_text='\n')
    tempDfs = [table.df for table in tables]
    tempDf = pd.concat(tempDfs, ignore_index=True)
    columns_list = tempDf.iloc[1]
    df1 = tempDf.rename(columns=columns_list).drop(tempDf.index[0]).reset_index(drop=True)
    df1 = df1.drop(df1.index[0]).reset_index(drop=True)
    df1 = df1[['Transaction Date', 'Value Date', 'Type of', 'Details', 'Instrument Id', 'Debits', 'Credits', 'Balance']]
    return df1

def extract_zyy_bank(file):
    # Save the uploaded file to a temporary fil

    # Extract account number and tables from the temporary file
    account_number, tables = extract_tables_from_pdf(file)
    if not tables:
        return None, None

    # Create a DataFrame from the extracted tables
    all_data = []
    for table in tables:
        all_data.extend(table)

    df = pd.DataFrame(all_data[1:], columns=all_data[0])
    return df, account_number

def extract_account_number(uploaded_file, patterns):
    if uploaded_file is None:
        return None
    
    with pdfplumber.open(io.BytesIO(uploaded_file.getvalue())) as pdf:
        full_text = ''
        for page in pdf.pages:
            full_text += page.extract_text() + '\n'
        
        if isinstance(patterns, list):
            for pattern in patterns:
                match = re.search(pattern, full_text, re.DOTALL)
                if match:
                    return str(match.group(1))
        else:
            match = re.search(patterns, full_text, re.DOTALL)
            if match:
                return str(match.group(1))
    return None

bank_patterns = {
    "Bank Muscat": [r'- Current Account\s+(\d{16})', r'Najahi - Current Account.*?\n.*?(\d{16})'],  
    "Bank Dhofar": r'Account No:\s+(\d{14})',  # 
    "OAB Bank": r'Account:\s*(\d+)'  # This pattern is already defined
}

def extract_tables_from_pdf(uploaded_file):
    account_number_pattern = r'Account:\s*(\d+)'  # Default pattern, adjust if needed
    account_number = extract_account_number(uploaded_file, account_number_pattern)
    
    with pdfplumber.open(io.BytesIO(uploaded_file.getvalue())) as pdf:
        all_tables = []
        for page in pdf.pages:
            tables = page.extract_tables()
            all_tables.extend(tables)

    return account_number, all_tables

def save_to_excel(account_number, tables):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Account Number'
    ws['B1'] = f"'{account_number}"
    current_row = 3
    for table in tables:
        for row in table:
            for col, cell in enumerate(row, start=1):
                ws.cell(row=current_row, column=col, value=cell)
            current_row += 1
        current_row += 1
    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)
    return excel_io

st.title("Bank Statement Extractor")

bank_option = st.sidebar.selectbox(
    "Select Bank",
    ["Bank Muscat", "Bank Dhofar", "OAB Bank"]
)

uploaded_file = st.file_uploader("Choose a PDF file", type="pdf")

if uploaded_file is not None:
    with st.spinner("Processing the PDF..."):
        if bank_option == "Bank Muscat":
            df = extract_xyz_bank(uploaded_file)
            patterns = bank_patterns[bank_option]
            account_number = extract_account_number(uploaded_file, patterns)
        elif bank_option == "Bank Dhofar":
            df = extract_yzx_bank(uploaded_file)
            pattern = bank_patterns[bank_option]
            account_number = extract_account_number(uploaded_file, pattern)
        elif bank_option == "OAB Bank":
            df, account_number = extract_zyy_bank(uploaded_file)

    if df is not None and not df.empty:
        if account_number:
            csv_filename = f'statement_{account_number}.csv'
            excel_io = save_to_excel(account_number, [df.values.tolist()])
            st.download_button(
                label="Download Excel File",
                data=excel_io,
                file_name=f'statement_{account_number}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            st.download_button(
                label="Download CSV File",
                data=df.to_csv(index=False),
                file_name=csv_filename,
                mime='text/csv'
            )
        else:
            st.write("Account Number not found.")
    else:
        st.write("No data to display.")
